#!/usr/bin/python

from PyQt4.QtCore import *
import xlrd
import xlwt
import openpyxl

import sys
import re

class SpreadSheetQuickSheet97(QObject):
    def __init__(self,h,parent = None):
        QObject.__init__(self,parent)
        self.h = h

    def name(self):
        return self.h.name
    
    def row_range(self):
        return (0,self.h.nrows - 1)
    
    def row_count(self):
        return self.h.nrows
    
    def col_range(self):
        return (0,self.h.ncols - 1)
 
    def col_count(self):
        return self.h.ncols
    
    def cell_value(self,row,col):
        v = self.h.cell_value(row,col)
        if v is None:
            return ''
        try:
            s = str(v)
        except:
            s = v.encode('utf-8')
        return s    
    def sync(self,diff):
        pass
    

        

class SpreadSheetQuickSheet07(QObject):
    def __init__(self,h,parent = None):
        QObject.__init__(self,parent)
        self.h = h

    def name(self):
        return self.h.title
    
    def row_range(self):
        return (0,len(self.h.rows) - 1)
    
    def row_count(self):
        return self.h.get_highest_row()
    
    def col_range(self):
        return (0,len(self.h.columns) - 1)

    def col_count(self):
        if self.h.get_highest_column() > 256:
            return 256
        return self.h.get_highest_column()
    
    def cell_value(self,xrow,xcol):
        
        cell = self.h.cell(row = xrow + 1,column = xcol + 1)
        if cell.value is None:
            return ''
        try:
            s = str(cell.value)
        except:
            s = cell.value.encode('utf-8')
        return s
    
    def sync(self,diff):
        for d in diff.keys():
            v = diff[d]
            index = str(d).split(',')
            self.h.cell(row = int(index[0]) + 1,column = int(index[1]) + 1).value = v     
        #print "Sheet %s , %d changes" % (self.name() , len(diff.keys()))
    
class SpreadSheetQuickSheet(QObject):

    rowCount = 50
    colCount = 24
    
    def __init__(self,h = None,parent=None):
        QObject.__init__(self,parent)
        self.h = h

    def name(self):
        return ''
    
    def row_range(self):
        return (0,self.rowCount - 1)
    
    def row_count(self):
        return self.rowCount
    
    def col_range(self):
        return (0,self.colCount - 1)

    def col_count(self):
        return self.colCount
    
    def cell_value(self,xrow,xcol):
        return ''
    
class SpreadSheetQuick(QObject):
    def __init__(self,fname = None,parent = None):
        QObject.__init__(self,parent)
        
        if not fname is None:
            self.fname = fname
            if re.match(r'.*\.xlsx$',self.fname.lower()):
                self.fmt = '2007'
                self.workbook = openpyxl.load_workbook(filename = fname)
            elif re.match(r'.*\.xls$',self.fname.lower()):
                self.fmt = '1997'
                self.workbook = xlrd.open_workbook(fname)
            else:
                self.fmt = ''
        else:
            self.fmt = 'default'
            
    def worksheets(self):
        if self.fmt == '2007':
            return self.workbook.get_sheet_names()
        elif self.fmt == '1997':
            return self.workbook.sheet_names()
        elif self.fmt == 'default':
            return ['Sheet1','Sheet2','Sheet3']
    
    def worksheet(self,name):
        if self.fmt == '2007':
            return SpreadSheetQuickSheet07(self.workbook.get_sheet_by_name(name),self)
        elif self.fmt == '1997':
            return SpreadSheetQuickSheet97(self.workbook.sheet_by_name(name),self)
        elif self.fmt == 'default':
            return SpreadSheetQuickSheet()
        
    def save_to_file(self,filename):
        if self.fmt == '2007':
            return self.workbook.save(filename)
        elif self.fmt == '1997':
            return False
        return False

    @staticmethod
    def create(xworkbook,filename):
        if re.match(r'.*\.xls$',filename.lower()):
            workbook = xlwt.Workbook()
            for xworksheet in xworkbook:
                sheet = workbook.add_sheet(xworksheet['name'])
                for d in xworksheet['data'].keys():
                    v = xworksheet['data'][d]
                    index = str(d).split(',')
                    sheet.write(int(index[0]),int(index[1]),v)
            try:
                workbook.save(filename)
                return True
            except:
                return False
            
        elif re.match(r'.*\.xlsx$',filename.lower()):
            workbook = openpyxl.Workbook()
            for i,xworksheet in enumerate(xworkbook):
                if i == 0:
                    sheet = workbook.active
                else:
                    sheet = workbook.create_sheet()
                sheet.title = xworksheet['name']
                for d in xworksheet['data'].keys():
                    v = xworksheet['data'][d]
                    index = str(d).split(',')
                    sheet.cell(row = int(index[0]) + 1,column = int(index[1]) + 1).value = v
            try:
                workbook.save(filename)
                return True
            except:
                return False            
        return False
    
    @staticmethod
    def save(xworkbook,workbook,filename):
        if re.match(r'.*\.xls$',filename.lower()):
            return SpreadSheetQuick.create(xworkbook,filename)
        elif re.match(r'.*\.xlsx$',filename.lower()):
            for xworksheet in xworkbook:
                #print "Sync %s" % xworksheet['name']
                sheet = workbook.worksheet(xworksheet['name'])
                sheet.sync(xworksheet['diff'])
            try:
                workbook.save_to_file(filename)
                return True
            except:
                return False                 
        return False    

def XlsHeader(i,j=None):
    RowStr = ''
    if i >= 0 and i <= 25:
        RowStr = chr(ord('A') + i)
    else:
        t = int(i / 26)
        RowStr = chr(ord('A') + t - 1) + chr(ord('A') + i - t * 26)
    if j is None:
        return RowStr
    else:
        return "%s%d" % (RowStr,(j+1))
